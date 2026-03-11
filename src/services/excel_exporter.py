import logging
from io import BytesIO
from openpyxl import Workbook
from src.storage.models import Record

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports financial records to an Excel file."""

    HEADERS = ["Date", "Type", "Category", "Description", "Amount"]

    def export(self, records: list[Record]) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Records"

        # Header row
        for col, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)

        # Data rows
        for i, record in enumerate(records, 2):
            ws.cell(row=i, column=1, value=record.created_at.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=i, column=2, value=record.entry_type)
            ws.cell(row=i, column=3, value=record.category)
            ws.cell(row=i, column=4, value=record.description)
            ws.cell(row=i, column=5, value=record.amount)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
