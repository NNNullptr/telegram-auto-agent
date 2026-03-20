import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.accounting.base import BaseAccounting
from src.storage.models import Transaction

logger = logging.getLogger(__name__)

EXCEL_DIR = Path(__file__).resolve().parent.parent.parent / "data"


class ExcelAccounting(BaseAccounting):
    """Excel file-backed accounting — appends transactions to a local .xlsx."""

    # [修复] 新增 CustomerName 列，与 SQLite/Notion 字段对齐
    HEADERS = ["Date", "Product", "Quantity", "UnitPrice", "TotalAmount", "CustomerName", "Description"]

    def __init__(self, export_dir: Path | None = None):
        self.export_dir = export_dir or EXCEL_DIR
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def _get_path(self, chat_id: int) -> Path:
        return self.export_dir / f"transactions_{chat_id}.xlsx"

    async def record_transaction(self, transaction: Transaction) -> str | None:
        try:
            path = self._get_path(transaction.chat_id)
            if path.exists():
                wb = load_workbook(path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Transactions"
                for col, header in enumerate(self.HEADERS, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = cell.font.copy(bold=True)

            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=transaction.created_at.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row, column=2, value=transaction.product)
            ws.cell(row=row, column=3, value=transaction.quantity)
            ws.cell(row=row, column=4, value=transaction.unit_price)
            ws.cell(row=row, column=5, value=transaction.total_amount)
            # [修复] 写入客户名称
            ws.cell(row=row, column=6, value=transaction.customer_name or "")
            ws.cell(row=row, column=7, value=transaction.description)

            wb.save(path)
            logger.info(f"Excel: appended transaction to {path}")
            return str(row - 1)
        except Exception as e:
            logger.error(f"Excel accounting failed: {e}")
            return None

    async def get_transactions(self, chat_id: int, limit: int = 50) -> list[Transaction]:
        path = self._get_path(chat_id)
        if not path.exists():
            return []
        try:
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            transactions = []
            for row in list(ws.iter_rows(min_row=2, values_only=True))[-limit:]:
                if row[0] is None:
                    continue
                transactions.append(Transaction(
                    chat_id=chat_id,
                    product=str(row[1] or ""),
                    quantity=int(row[2] or 1),
                    unit_price=float(row[3] or 0),
                    total_amount=float(row[4] or 0),
                    # [修复] 读取客户名称（兼容旧文件：列可能不存在）
                    customer_name=str(row[5] or "") if len(row) > 6 else "",
                    description=str(row[6] or "") if len(row) > 6 else str(row[5] or ""),
                ))
            return transactions
        except Exception as e:
            logger.error(f"Excel get_transactions failed: {e}")
            return []
