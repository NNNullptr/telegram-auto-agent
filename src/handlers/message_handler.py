#引入各种模块
import logging
from io import BytesIO
from telegram import Update
from telegram.ext import ContextTypes
from openpyxl import Workbook

from config import settings
from src.agents.classifier import ClassifierAgent
from src.agents.consulting import ConsultingAgent
from src.agents.chatting import ChattingAgent
from src.agents.purchasing import PurchasingAgent
from src.llm import LLMFactory
from src.accounting.sqlite_accounting import SQLiteAccounting
from src.accounting.notion_accounting import NotionAccounting
from src.accounting.excel_accounting import ExcelAccounting
from src.accounting.composite import CompositeAccounting
from src.services.notion_client import NotionService
from src.services.excel_exporter import ExcelExporter
from src.storage.database import Database
from src.storage.models import Transaction
from src.utils.context_manager import ContextManager
from src.graph.graph import build_graph
from src.graph.nodes import is_manual_mode, set_manual_mode

logger = logging.getLogger(__name__)

# 确认关键词：用户回复这些词时视为同意下单（和 purchasing.py 保持一致）
# [修复] 使用精确匹配（整条消息 == 关键词），防止 "确定不要" 误触发确认
_CONFIRM_KEYWORDS = {"确认", "是的", "对", "好的", "确定", "yes", "confirm", "ok", "sure", "y"}

# 取消关键词：用户回复这些词时视为放弃待确认订单
_CANCEL_KEYWORDS = {"取消", "不要", "不用", "算了", "不下了", "cancel", "no", "nope", "n"}


class MessageHandler:
    """中央消息分发器：使用 LangGraph 状态机进行路由"""
    def __init__(self):
        llm = LLMFactory.create()

        self.classifier = ClassifierAgent(llm)
        self.consulting = ConsultingAgent(llm)
        self.chatting = ChattingAgent(llm)
        self.purchasing = PurchasingAgent(llm)

        self.db = Database()
        self.notion = NotionService()
        self.excel_exporter = ExcelExporter()
        self.context = ContextManager()
        self.accounting = self._build_accounting()

        self.graph = build_graph(
            classifier=self.classifier,
            consulting_agent=self.consulting,
            chatting_agent=self.chatting,
            purchasing_agent=self.purchasing,
            accounting=self.accounting,
            admin_chat_id=settings.admin_id or None,
            confidence_threshold=settings.confidence_threshold,
        )

        # message_id -> user_chat_id 映射，用于 admin Reply 时找到原用户
        self.forwarded_map: dict[int, int] = {}

    def _build_accounting(self) -> CompositeAccounting:
        backends = []
        enabled = [b.strip() for b in settings.accounting_backends.split(",")]
        if "sqlite" in enabled:
            backends.append(SQLiteAccounting(self.db))
        if "notion" in enabled:
            backends.append(NotionAccounting(self.notion))
        if "excel" in enabled:
            backends.append(ExcelAccounting())
        if not backends:
            backends.append(SQLiteAccounting(self.db))

        # [新增] 配置冲突检测：NOTION_ENABLED=true 但 ACCOUNTING_BACKENDS 中没有 notion
        # 这是常见的配置遗漏，原来完全静默导致极难排查
        if settings.notion_enabled and "notion" not in enabled:
            logger.warning(
                "⚠️ NOTION_ENABLED=true 但 ACCOUNTING_BACKENDS 中未包含 'notion'，"
                "订单将不会同步到 Notion！请在 .env 中设置 ACCOUNTING_BACKENDS=sqlite,notion"
            )

        logger.info(f"Accounting backends: {[type(b).__name__ for b in backends]}")
        return CompositeAccounting(backends)

    async def init(self) -> None:
        await self.db.init()

        # 恢复 manual 模式状态
        modes = await self.db.load_manual_modes()
        for chat_id, enabled in modes.items():
            set_manual_mode(chat_id, enabled)
        if modes:
            logger.info(f"Restored {len(modes)} manual mode sessions")

        # [修复] 恢复待确认订单：根据 DB 中的 created_at 计算正确的剩余 TTL
        # 原实现直接调用 set_pending_order → 获得全新的 30 分钟 TTL，
        # 导致已过期的订单在重启后被错误复活
        from datetime import datetime as _dt
        pending_orders = await self.db.load_pending_orders()
        expired_count = 0
        for chat_id, (order, created_at_iso) in pending_orders.items():
            try:
                created_at = _dt.fromisoformat(created_at_iso)
                age_seconds = (_dt.now() - created_at).total_seconds()
                remaining = self.context.PENDING_ORDER_TTL - age_seconds
                if remaining <= 0:
                    # 已过期，从 DB 中也清理掉
                    await self.db.delete_pending_order(chat_id)
                    expired_count += 1
                    continue
                # 按正确的剩余时间设置过期
                import time as _time
                self.context.set_pending_order(chat_id, order, expires_at=_time.time() + remaining)
            except (ValueError, TypeError) as e:
                logger.warning(f"Skipping pending order for {chat_id}: bad created_at: {e}")
                await self.db.delete_pending_order(chat_id)
        restored = len(pending_orders) - expired_count
        if restored:
            logger.info(f"Restored {restored} pending orders")
        if expired_count:
            logger.info(f"Cleaned up {expired_count} expired pending orders from DB")

    async def handle_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if not update.message or not update.message.text:
            return

        chat_id = update.effective_chat.id
        user_message = update.message.text.strip()
        logger.info(f"[{chat_id}] Received: {user_message}")

        # ── Manual mode: 转发给 admin ──
        if is_manual_mode(chat_id) and settings.admin_id:
            await self._forward_to_admin(update, chat_id, user_message, context)
            return

        # ── 待确认订单状态机（优先于 classifier，避免误分类） ──
        # 先检查取消，再检查确认，两者均无需调用 LangGraph 节省开销
        if self._check_cancel_order(chat_id, user_message):
            await self._cancel_order(update, context, chat_id)
            return

        pending_order = self._check_pending_order(chat_id, user_message)
        if pending_order:
            await self._confirm_order(update, context, chat_id, pending_order)
            return

        # ── 正常流程：invoke LangGraph ──
        # history 仅在进入 LangGraph 时才需要，延迟到此处获取
        history = self.context.get_history(chat_id)
        state = {
            "chat_id": chat_id,
            "user_message": user_message,
            "history": history,
            "is_manual": False,
            "intent": None,
            "confidence": 0.0,
            "response": "",
            "extracted_order": None,
            "order_confirmed": False,
        }

        result = await self.graph.ainvoke(state)
        reply = result.get("response", "Sorry, something went wrong.")
        
        extracted_order = result.get("extracted_order")
        is_order_confirmed = result.get("order_confirmed")

        # ── 状态机逻辑 ──
        # 场景 A: AI 解析出订单但尚未确认，存入状态机等待用户回复
        if extracted_order and not is_order_confirmed:
            self.context.set_pending_order(chat_id, extracted_order)
            # [新增] 同步写入数据库，防止 bot 重启后订单丢失
            await self.db.save_pending_order(chat_id, extracted_order)
            logger.info(f"[{chat_id}] 订单存入待确认状态（已持久化）")

        # 场景 B: 订单在 Graph 内部被直接确认（安全兜底，一般不触发）
        elif is_order_confirmed and extracted_order:
            self.context.clear_pending_order(chat_id)
            # [新增] 同步删除数据库记录
            await self.db.delete_pending_order(chat_id)
            await self._notify_admin_and_enter_manual(chat_id, extracted_order, context)

        # 如果自动升级到 manual，通知 admin
        if result.get("intent") == "manual" and settings.admin_id:
            try:
                await context.bot.send_message(
                    chat_id=settings.admin_id,
                    text=(
                        f"🔔 用户 {chat_id} 请求人工帮助\n"
                        f"消息：{user_message}\n"
                        f"使用 /takeover {chat_id} 接管\n"
                        f"使用 /release {chat_id} 释放"
                    ),
                )
            except Exception as e:
                logger.error(f"Failed to notify admin: {e}")

        self.context.add_message(chat_id, "user", user_message)
        self.context.add_message(chat_id, "assistant", reply)

        await update.message.reply_text(reply)
        logger.info(f"[{chat_id}] Intent: {result.get('intent')} | Replied: {reply[:80]}...")

    def _check_pending_order(self, chat_id: int, user_message: str) -> dict | None:
        """检查用户是否在回复确认关键词，且当前存在待确认订单。

        直接从 ContextManager 获取结构化订单字典，不再做文本解析。
        过期订单由 ContextManager 惰性清理，此处无需处理。
        """
        pending_order = self.context.get_pending_order(chat_id)
        if not pending_order:
            return None
        # [修复] 精确匹配：整条消息必须是关键词本身，而非子串包含
        # 防止 "确定不要"、"我不确定" 等复合语句误触发
        msg_lower = user_message.strip().lower()
        if msg_lower in _CONFIRM_KEYWORDS:
            return pending_order
        return None

    def _check_cancel_order(self, chat_id: int, user_message: str) -> bool:
        """检查用户是否在回复取消关键词，且当前存在待确认订单。

        取消检查优先于确认检查执行，防止模糊词义被误判为确认。
        """
        if not self.context.get_pending_order(chat_id):
            return False
        # [修复] 同样使用精确匹配
        msg_lower = user_message.strip().lower()
        return msg_lower in _CANCEL_KEYWORDS

    async def _cancel_order(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        chat_id: int,
    ) -> None:
        """处理订单取消：清除内存状态 → 删除 DB 记录 → 回复用户。"""
        self.context.clear_pending_order(chat_id)
        # 同步删除数据库中的待确认记录
        await self.db.delete_pending_order(chat_id)

        reply = "好的，已取消本次订单。如需重新下单，请随时告诉我。"
        self.context.add_message(chat_id, "user", update.message.text)
        self.context.add_message(chat_id, "assistant", reply)
        await update.message.reply_text(reply)
        logger.info(f"[{chat_id}] 用户主动取消待确认订单")

    async def _confirm_order(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        chat_id: int,
        order: dict,
    ) -> None:
        """处理订单确认：记录交易 → 通知 admin → 进入 manual 模式。

        [修改] Transaction 新增 customer_name 字段。
        """
        # 1. 记录交易到数据库
        transaction = Transaction(
            chat_id=chat_id,
            product=order.get("product", "unknown"),
            quantity=order.get("quantity", 1),
            unit_price=order.get("unit_price", 0),
            total_amount=order.get("total_amount", 0),
            customer_name=order.get("customer_name", ""),
            # [修复] 补上 description，原实现丢失了 LLM 提取的订单备注
            description=order.get("description", ""),
        )
        record_id = await self.accounting.record_transaction(transaction)
        logger.info(f"[{chat_id}] Transaction recorded: {record_id}")

        # 2. 回复用户（[修改] 增加客户名称显示）
        customer = order.get("customer_name", "")
        # [修复] 使用 .get() 防止 KeyError
        reply = (
            f"✅ 订单已确认！\n"
            + (f"👤 客户：{customer}\n" if customer else "")
            + f"📦 {order.get('product', '未知商品')} x{order.get('quantity', 1)}\n"
            f"💰 总计 ¥{order.get('total_amount', 0):.2f}\n"
            f"正在为您转接人工客服确认详情..."
        )
        self.context.add_message(chat_id, "user", update.message.text)
        self.context.add_message(chat_id, "assistant", reply)
        await update.message.reply_text(reply)

        self.context.clear_pending_order(chat_id)
        # [新增] 订单确认后同步删除数据库中的待确认记录
        await self.db.delete_pending_order(chat_id)

        # 3. 通知 admin + 进入 manual 模式
        await self._notify_admin_and_enter_manual(chat_id, order, context)

    async def _notify_admin_and_enter_manual(
        self,
        chat_id: int,
        order: dict,
        context: ContextTypes.DEFAULT_TYPE,
    ) -> None:
        """Notify admin about a confirmed order and enter manual mode.

        [修改] 通知消息中增加客户名称显示。
        """
        if settings.admin_id:
            customer = order.get("customer_name", "未提供")
            order_text = (
                f"🛒 新订单！\n"
                f"👤 客户名称：{customer}\n"
                f"🆔 用户 ID：{chat_id}\n"
                f"📦 商品：{order.get('product', '?')}\n"
                f"🔢 数量：{order.get('quantity', '?')}\n"
                f"💰 单价：¥{order.get('unit_price', 0):.2f}\n"
                f"💵 总计：¥{order.get('total_amount', 0):.2f}\n\n"
                f"用户已自动进入人工模式。\n"
                f"用户后续消息会转发给您，直接 Reply 即可回复。\n"
                f"完成后发送 /release {chat_id} 切回 AI。"
            )
            try:
                await context.bot.send_message(
                    chat_id=settings.admin_id, text=order_text
                )
            except Exception as e:
                logger.error(f"Failed to send order to admin: {e}")

        set_manual_mode(chat_id, True)
        await self.db.save_manual_mode(chat_id, True)
        logger.info(f"[{chat_id}] Auto-entered manual mode after order")

    async def _forward_to_admin(
        self,
        update: Update,
        chat_id: int,
        user_message: str,
        context: ContextTypes.DEFAULT_TYPE,
    ) -> None:
        """manual 模式下转发用户消息给 admin。

        [修复] 用户消息仍写入上下文（admin_handler 中 admin 回复也会写入），
        保持对话历史完整对称。release 时会统一清空，所以不会污染后续 AI 对话。
        """
        text = f"💬 [用户 {chat_id}]\n{user_message}"
        try:
            sent = await context.bot.send_message(
                chat_id=settings.admin_id, text=text
            )
            self.forwarded_map[sent.message_id] = chat_id
            # [修复] 每次插入后若超过上限，只移除最旧的一条（FIFO）
            if len(self.forwarded_map) > 500:
                oldest_key = next(iter(self.forwarded_map))
                del self.forwarded_map[oldest_key]
        except Exception as e:
            logger.error(f"Failed to forward to admin: {e}")

        self.context.add_message(chat_id, "user", user_message)
        await update.message.reply_text("已收到，正在等待人工回复...")

    async def handle_export(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        """Handle /export — 仅管理员可用，导出指定客户的交易记录。

        用法：/export <chat_id>
        [修改] 表格新增"客户名称"列。
        """
        if update.effective_user.id != settings.admin_id:
            await update.message.reply_text("仅管理员可导出账单。")
            return

        if not context.args:
            await update.message.reply_text("用法：/export <chat_id>")
            return

        try:
            target_chat_id = int(context.args[0])
        except ValueError:
            await update.message.reply_text("chat_id 无效。")
            return

        transactions = await self.accounting.get_transactions(target_chat_id)
        if not transactions:
            await update.message.reply_text("暂无交易记录。")
            return

        output = self._build_excel(transactions, include_chat_id=False)
        await update.message.reply_document(
            document=output,
            filename=f"transactions_{target_chat_id}.xlsx",
            caption=f"已导出 {len(transactions)} 条交易记录。",
        )

    # [新增] /exportall 命令：导出所有客户的交易记录，表格增加"客户ID"列用于区分
    async def handle_export_all(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        """Handle /exportall — 仅管理员可用，一次导出所有客户的交易记录。

        导出的 Excel 表格包含"客户ID"列，方便区分不同客户。
        """
        if update.effective_user.id != settings.admin_id:
            await update.message.reply_text("仅管理员可导出账单。")
            return

        # [说明] 从 SQLiteAccounting 获取全部交易记录
        sqlite_backend = None
        for backend in self.accounting.backends:
            if isinstance(backend, SQLiteAccounting):
                sqlite_backend = backend
                break

        if not sqlite_backend:
            await update.message.reply_text("SQLite 后端未启用，无法导出。")
            return

        transactions = await sqlite_backend.get_all_transactions()
        if not transactions:
            await update.message.reply_text("暂无任何交易记录。")
            return

        # [说明] include_chat_id=True 会在表格第一列插入"客户ID"
        output = self._build_excel(transactions, include_chat_id=True)
        await update.message.reply_document(
            document=output,
            filename="all_transactions.xlsx",
            caption=f"已导出全部 {len(transactions)} 条交易记录。",
        )

    @staticmethod
    def _build_excel(transactions: list[Transaction], include_chat_id: bool = False) -> BytesIO:
        """[新增] 通用 Excel 生成工具方法。

        Args:
            transactions: 交易记录列表
            include_chat_id: 是否在第一列插入"客户ID"（/exportall 时为 True）
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # [修改] 表头增加"客户名称"列；include_chat_id 时增加"客户ID"列
        if include_chat_id:
            headers = ["客户ID", "客户名称", "日期", "商品", "数量", "单价", "总计", "备注"]
        else:
            headers = ["客户名称", "日期", "商品", "数量", "单价", "总计", "备注"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = cell.font.copy(bold=True)

        for i, t in enumerate(transactions, 2):
            col = 1
            if include_chat_id:
                ws.cell(row=i, column=col, value=t.chat_id)
                col += 1
            ws.cell(row=i, column=col, value=t.customer_name or "")
            col += 1
            ws.cell(row=i, column=col, value=t.created_at.strftime("%Y-%m-%d %H:%M"))
            col += 1
            ws.cell(row=i, column=col, value=t.product)
            col += 1
            ws.cell(row=i, column=col, value=t.quantity)
            col += 1
            ws.cell(row=i, column=col, value=t.unit_price)
            col += 1
            ws.cell(row=i, column=col, value=t.total_amount)
            col += 1
            ws.cell(row=i, column=col, value=t.description)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def handle_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        await update.message.reply_text(
            "👋 Welcome!\n\n"
            "• 💬 咨询商品 — 直接提问\n"
            "• 🛒 购买下单 — 说「我要买...」\n"
            "• 💬 闲聊 — 随便聊\n"
            "• 👨‍💼 转人工 — 说「转人工」\n"
            "• 📊 导出账单 — /export <chat_id>\n"
            "• 📊 导出全部 — /exportall"
        )
